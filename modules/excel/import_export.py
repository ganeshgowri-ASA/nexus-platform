"""Import and export functionality for spreadsheets."""
from typing import Optional, Dict, Any, List
import pandas as pd
from io import BytesIO
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


class ImportExport:
    """Import and export spreadsheet data in various formats."""

    @staticmethod
    def import_excel(file_path: str, sheet_name: Optional[str] = None) -> pd.DataFrame:
        """
        Import Excel file.

        Args:
            file_path: Path to Excel file
            sheet_name: Sheet name to import (None = first sheet)

        Returns:
            DataFrame with imported data
        """
        return pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')

    @staticmethod
    def import_excel_bytes(file_bytes: bytes, sheet_name: Optional[str] = None) -> pd.DataFrame:
        """
        Import Excel from bytes.

        Args:
            file_bytes: Excel file as bytes
            sheet_name: Sheet name to import

        Returns:
            DataFrame with imported data
        """
        buffer = BytesIO(file_bytes)
        return pd.read_excel(buffer, sheet_name=sheet_name, engine='openpyxl')

    @staticmethod
    def import_csv(file_path: str, delimiter: str = ',',
                  encoding: str = 'utf-8') -> pd.DataFrame:
        """
        Import CSV file.

        Args:
            file_path: Path to CSV file
            delimiter: CSV delimiter
            encoding: File encoding

        Returns:
            DataFrame with imported data
        """
        return pd.read_csv(file_path, delimiter=delimiter, encoding=encoding)

    @staticmethod
    def import_csv_bytes(file_bytes: bytes, delimiter: str = ',',
                        encoding: str = 'utf-8') -> pd.DataFrame:
        """
        Import CSV from bytes.

        Args:
            file_bytes: CSV file as bytes
            delimiter: CSV delimiter
            encoding: File encoding

        Returns:
            DataFrame with imported data
        """
        buffer = BytesIO(file_bytes)
        return pd.read_csv(buffer, delimiter=delimiter, encoding=encoding)

    @staticmethod
    def import_json(file_path: str) -> pd.DataFrame:
        """
        Import JSON file.

        Args:
            file_path: Path to JSON file

        Returns:
            DataFrame with imported data
        """
        return pd.read_json(file_path)

    @staticmethod
    def import_json_bytes(file_bytes: bytes) -> pd.DataFrame:
        """
        Import JSON from bytes.

        Args:
            file_bytes: JSON file as bytes

        Returns:
            DataFrame with imported data
        """
        buffer = BytesIO(file_bytes)
        return pd.read_json(buffer)

    @staticmethod
    def export_excel(df: pd.DataFrame, file_path: str,
                    sheet_name: str = 'Sheet1',
                    with_formatting: bool = False,
                    cell_styles: Optional[Dict] = None) -> None:
        """
        Export to Excel file.

        Args:
            df: DataFrame to export
            file_path: Output file path
            sheet_name: Sheet name
            with_formatting: Whether to apply formatting
            cell_styles: Optional cell styles dict
        """
        if with_formatting and cell_styles:
            # Use openpyxl for formatted export
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name

            # Write data
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)

                    # Apply styles if provided
                    if cell_styles and (r_idx, c_idx) in cell_styles:
                        style = cell_styles[(r_idx, c_idx)]
                        if 'font' in style:
                            cell.font = Font(**style['font'])
                        if 'fill' in style:
                            cell.fill = PatternFill(**style['fill'])
                        if 'alignment' in style:
                            cell.alignment = Alignment(**style['alignment'])

            wb.save(file_path)
        else:
            # Simple export without formatting
            df.to_excel(file_path, sheet_name=sheet_name, index=False, engine='openpyxl')

    @staticmethod
    def export_excel_bytes(df: pd.DataFrame, sheet_name: str = 'Sheet1') -> bytes:
        """
        Export to Excel bytes.

        Args:
            df: DataFrame to export
            sheet_name: Sheet name

        Returns:
            Excel file as bytes
        """
        buffer = BytesIO()
        df.to_excel(buffer, sheet_name=sheet_name, index=False, engine='openpyxl')
        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def export_csv(df: pd.DataFrame, file_path: str,
                  delimiter: str = ',', encoding: str = 'utf-8') -> None:
        """
        Export to CSV file.

        Args:
            df: DataFrame to export
            file_path: Output file path
            delimiter: CSV delimiter
            encoding: File encoding
        """
        df.to_csv(file_path, sep=delimiter, encoding=encoding, index=False)

    @staticmethod
    def export_csv_bytes(df: pd.DataFrame, delimiter: str = ',') -> bytes:
        """
        Export to CSV bytes.

        Args:
            df: DataFrame to export
            delimiter: CSV delimiter

        Returns:
            CSV file as bytes
        """
        return df.to_csv(sep=delimiter, index=False).encode('utf-8')

    @staticmethod
    def export_json(df: pd.DataFrame, file_path: str,
                   orient: str = 'records', indent: int = 2) -> None:
        """
        Export to JSON file.

        Args:
            df: DataFrame to export
            file_path: Output file path
            orient: JSON orientation ('records', 'index', 'columns', etc.)
            indent: JSON indentation
        """
        df.to_json(file_path, orient=orient, indent=indent)

    @staticmethod
    def export_json_bytes(df: pd.DataFrame, orient: str = 'records',
                         indent: int = 2) -> bytes:
        """
        Export to JSON bytes.

        Args:
            df: DataFrame to export
            orient: JSON orientation
            indent: JSON indentation

        Returns:
            JSON file as bytes
        """
        return df.to_json(orient=orient, indent=indent).encode('utf-8')

    @staticmethod
    def export_html(df: pd.DataFrame, file_path: str,
                   with_styles: bool = True) -> None:
        """
        Export to HTML file.

        Args:
            df: DataFrame to export
            file_path: Output file path
            with_styles: Whether to include styling
        """
        html = df.to_html(index=False, escape=False)

        if with_styles:
            html = f"""
<!DOCTYPE html>
<html>
<head>
    <style>
        table {{
            border-collapse: collapse;
            width: 100%;
            font-family: Arial, sans-serif;
        }}
        th, td {{
            border: 1px solid #ddd;
            padding: 8px;
            text-align: left;
        }}
        th {{
            background-color: #4CAF50;
            color: white;
        }}
        tr:nth-child(even) {{
            background-color: #f2f2f2;
        }}
    </style>
</head>
<body>
    {html}
</body>
</html>
"""

        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(html)

    @staticmethod
    def export_pdf(df: pd.DataFrame, file_path: str) -> None:
        """
        Export to PDF file.

        Args:
            df: DataFrame to export
            file_path: Output file path

        Note:
            Requires additional dependencies (reportlab or weasyprint)
        """
        # Create HTML first
        html_path = file_path.replace('.pdf', '.html')
        ImportExport.export_html(df, html_path)

        # Convert to PDF (simplified - in production use proper PDF library)
        # This is a placeholder - implement with reportlab or weasyprint
        print(f"PDF export requires additional dependencies. HTML saved to {html_path}")

    @staticmethod
    def import_from_google_sheets(sheet_id: str, credentials_path: str) -> pd.DataFrame:
        """
        Import from Google Sheets.

        Args:
            sheet_id: Google Sheets ID
            credentials_path: Path to credentials JSON

        Returns:
            DataFrame with imported data

        Note:
            Requires gspread and google-auth packages
        """
        # Placeholder - implement with gspread
        raise NotImplementedError("Google Sheets import requires gspread library")

    @staticmethod
    def export_to_google_sheets(df: pd.DataFrame, sheet_id: str,
                               credentials_path: str) -> None:
        """
        Export to Google Sheets.

        Args:
            df: DataFrame to export
            sheet_id: Google Sheets ID
            credentials_path: Path to credentials JSON

        Note:
            Requires gspread and google-auth packages
        """
        # Placeholder - implement with gspread
        raise NotImplementedError("Google Sheets export requires gspread library")

    @staticmethod
    def detect_format(file_path: str) -> str:
        """
        Detect file format from extension.

        Args:
            file_path: File path

        Returns:
            Format string ('excel', 'csv', 'json', etc.)
        """
        ext = file_path.lower().split('.')[-1]

        format_map = {
            'xlsx': 'excel',
            'xls': 'excel',
            'csv': 'csv',
            'tsv': 'tsv',
            'json': 'json',
            'html': 'html',
            'pdf': 'pdf'
        }

        return format_map.get(ext, 'unknown')

    @staticmethod
    def auto_import(file_path: str) -> pd.DataFrame:
        """
        Auto-detect format and import.

        Args:
            file_path: File path

        Returns:
            DataFrame with imported data
        """
        format_type = ImportExport.detect_format(file_path)

        if format_type == 'excel':
            return ImportExport.import_excel(file_path)
        elif format_type == 'csv':
            return ImportExport.import_csv(file_path)
        elif format_type == 'tsv':
            return ImportExport.import_csv(file_path, delimiter='\t')
        elif format_type == 'json':
            return ImportExport.import_json(file_path)
        else:
            raise ValueError(f"Unsupported file format: {format_type}")
