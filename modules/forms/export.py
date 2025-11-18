"""
Export Module

Handles exporting form responses to various formats: CSV, Excel, PDF, JSON.
Includes Google Sheets integration capabilities.
"""

import csv
import json
from typing import Any, Dict, List, Optional
from datetime import datetime
import io


class DataExporter:
    """Main exporter class for form responses"""

    def __init__(self, responses: List[Any], form_fields: List[Any]):
        """
        Initialize exporter

        Args:
            responses: List of FormResponse objects
            form_fields: List of Field objects
        """
        self.responses = responses
        self.form_fields = form_fields

    def export_to_csv(self, filepath: Optional[str] = None,
                     include_metadata: bool = True) -> str:
        """
        Export responses to CSV format

        Args:
            filepath: Path to save file, or None to return string
            include_metadata: Include response metadata columns

        Returns:
            CSV content as string
        """
        # Prepare headers
        headers = self._get_headers(include_metadata)

        # Prepare rows
        rows = []
        for response in self.responses:
            row = self._response_to_row(response, include_metadata)
            rows.append(row)

        # Write to CSV
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerows(rows)

        csv_content = output.getvalue()
        output.close()

        # Save to file if filepath provided
        if filepath:
            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                f.write(csv_content)

        return csv_content

    def export_to_json(self, filepath: Optional[str] = None,
                      pretty: bool = True) -> str:
        """
        Export responses to JSON format

        Args:
            filepath: Path to save file, or None to return string
            pretty: Pretty print JSON

        Returns:
            JSON content as string
        """
        data = {
            "form_id": self.responses[0].form_id if self.responses else None,
            "export_date": datetime.now().isoformat(),
            "total_responses": len(self.responses),
            "fields": [
                {
                    "id": field.id,
                    "label": field.label,
                    "type": field.field_type.value,
                }
                for field in self.form_fields
            ],
            "responses": [response.to_dict() for response in self.responses],
        }

        indent = 2 if pretty else None
        json_content = json.dumps(data, indent=indent, default=str)

        # Save to file if filepath provided
        if filepath:
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(json_content)

        return json_content

    def export_to_excel(self, filepath: str,
                       include_metadata: bool = True) -> bool:
        """
        Export responses to Excel format

        Args:
            filepath: Path to save Excel file
            include_metadata: Include response metadata columns

        Returns:
            True if successful
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            print("openpyxl not installed. Install with: pip install openpyxl")
            return False

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Form Responses"

        # Add headers
        headers = self._get_headers(include_metadata)
        ws.append(headers)

        # Style header row
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add data rows
        for response in self.responses:
            row = self._response_to_row(response, include_metadata)
            ws.append(row)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Add summary sheet
        summary_ws = wb.create_sheet("Summary")
        self._add_summary_sheet(summary_ws)

        # Save workbook
        wb.save(filepath)
        return True

    def export_to_pdf(self, filepath: str) -> bool:
        """
        Export responses to PDF format

        Args:
            filepath: Path to save PDF file

        Returns:
            True if successful
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        except ImportError:
            print("reportlab not installed. Install with: pip install reportlab")
            return False

        # Create PDF document
        doc = SimpleDocTemplate(filepath, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#4F46E5'),
            spaceAfter=30,
        )
        elements.append(Paragraph("Form Responses Report", title_style))
        elements.append(Spacer(1, 0.2 * inch))

        # Summary information
        summary_data = [
            ["Total Responses", str(len(self.responses))],
            ["Export Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Total Fields", str(len(self.form_fields))],
        ]

        summary_table = Table(summary_data, colWidths=[2 * inch, 3 * inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.grey),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('BACKGROUND', (1, 0), (1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 0.5 * inch))

        # Individual responses
        for i, response in enumerate(self.responses, 1):
            # Response header
            response_title = Paragraph(f"Response #{i}", styles['Heading2'])
            elements.append(response_title)
            elements.append(Spacer(1, 0.1 * inch))

            # Response metadata
            metadata = [
                ["Submitted", response.submitted_at.strftime("%Y-%m-%d %H:%M:%S")],
                ["Time Spent", f"{response.time_spent}s"],
            ]

            if response.respondent_email:
                metadata.append(["Email", response.respondent_email])

            metadata_table = Table(metadata, colWidths=[1.5 * inch, 3.5 * inch])
            metadata_table.setStyle(TableStyle([
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('TEXTCOLOR', (0, 0), (0, -1), colors.grey),
            ]))
            elements.append(metadata_table)
            elements.append(Spacer(1, 0.2 * inch))

            # Response data
            response_data = []
            for field in self.form_fields:
                value = response.data.get(field.id, "")
                if isinstance(value, list):
                    value = ", ".join(str(v) for v in value)
                response_data.append([field.label, str(value)])

            data_table = Table(response_data, colWidths=[2 * inch, 4 * inch])
            data_table.setStyle(TableStyle([
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            elements.append(data_table)

            # Page break between responses (except last)
            if i < len(self.responses):
                elements.append(PageBreak())

        # Build PDF
        doc.build(elements)
        return True

    def export_to_google_sheets(self, spreadsheet_id: str,
                               credentials_path: str) -> bool:
        """
        Export responses to Google Sheets

        Args:
            spreadsheet_id: Google Sheets spreadsheet ID
            credentials_path: Path to Google API credentials JSON

        Returns:
            True if successful
        """
        try:
            from google.oauth2.service_account import Credentials
            from googleapiclient.discovery import build
        except ImportError:
            print("Google API client not installed. Install with: pip install google-auth google-api-python-client")
            return False

        try:
            # Set up credentials
            SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
            creds = Credentials.from_service_account_file(credentials_path, scopes=SCOPES)

            # Build service
            service = build('sheets', 'v4', credentials=creds)

            # Prepare data
            headers = self._get_headers(include_metadata=True)
            rows = [self._response_to_row(r, include_metadata=True) for r in self.responses]
            data = [headers] + rows

            # Update spreadsheet
            body = {'values': data}
            service.spreadsheets().values().update(
                spreadsheetId=spreadsheet_id,
                range='A1',
                valueInputOption='RAW',
                body=body
            ).execute()

            return True

        except Exception as e:
            print(f"Error exporting to Google Sheets: {e}")
            return False

    def export_summary_statistics(self, filepath: str) -> bool:
        """
        Export summary statistics to JSON

        Args:
            filepath: Path to save file

        Returns:
            True if successful
        """
        from .analytics import FormAnalytics

        analytics = FormAnalytics(self.responses, self.form_fields)
        summary = analytics.export_summary_report()

        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(summary, f, indent=2, default=str)

        return True

    def _get_headers(self, include_metadata: bool) -> List[str]:
        """Get CSV/Excel headers"""
        headers = []

        if include_metadata:
            headers.extend([
                "Response ID",
                "Submitted At",
                "Time Spent (s)",
                "Email",
                "Complete",
            ])

        # Add field labels
        headers.extend([field.label for field in self.form_fields])

        return headers

    def _response_to_row(self, response: Any, include_metadata: bool) -> List[Any]:
        """Convert response to row data"""
        row = []

        if include_metadata:
            row.extend([
                response.id,
                response.submitted_at.strftime("%Y-%m-%d %H:%M:%S"),
                response.time_spent,
                response.respondent_email or "",
                "Yes" if response.is_complete else "No",
            ])

        # Add field values
        for field in self.form_fields:
            value = response.data.get(field.id, "")
            # Convert lists to comma-separated strings
            if isinstance(value, list):
                value = ", ".join(str(v) for v in value)
            row.append(value)

        return row

    def _add_summary_sheet(self, worksheet: Any) -> None:
        """Add summary statistics to Excel worksheet"""
        from .analytics import FormAnalytics

        analytics = FormAnalytics(self.responses, self.form_fields)
        metrics = analytics.get_overview_metrics()

        worksheet.append(["Form Response Summary"])
        worksheet.append([])
        worksheet.append(["Metric", "Value"])
        worksheet.append(["Total Submissions", metrics.total_submissions])
        worksheet.append(["Completion Rate", f"{metrics.completion_rate:.2f}%"])
        worksheet.append(["Average Time", f"{metrics.average_time:.0f}s"])
        worksheet.append(["Median Time", f"{metrics.median_time:.0f}s"])
        worksheet.append(["Abandonment Rate", f"{metrics.abandonment_rate:.2f}%"])

        # Style summary sheet
        try:
            from openpyxl.styles import Font, PatternFill
            worksheet['A1'].font = Font(size=16, bold=True)
            worksheet['A3'].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            worksheet['B3'].fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        except:
            pass


class BatchExporter:
    """Export multiple forms at once"""

    def __init__(self):
        self.forms_data: List[Dict[str, Any]] = []

    def add_form(self, form_id: str, responses: List[Any], fields: List[Any]) -> None:
        """Add a form to batch export"""
        self.forms_data.append({
            "form_id": form_id,
            "responses": responses,
            "fields": fields,
        })

    def export_all_to_zip(self, output_path: str, format: str = "csv") -> bool:
        """
        Export all forms to a ZIP file

        Args:
            output_path: Path to save ZIP file
            format: Export format (csv, json, excel)

        Returns:
            True if successful
        """
        try:
            import zipfile
        except ImportError:
            print("zipfile module not available")
            return False

        try:
            with zipfile.ZipFile(output_path, 'w') as zipf:
                for form_data in self.forms_data:
                    form_id = form_data["form_id"]
                    exporter = DataExporter(form_data["responses"], form_data["fields"])

                    if format == "csv":
                        content = exporter.export_to_csv()
                        filename = f"form_{form_id}.csv"
                        zipf.writestr(filename, content)

                    elif format == "json":
                        content = exporter.export_to_json()
                        filename = f"form_{form_id}.json"
                        zipf.writestr(filename, content)

            return True

        except Exception as e:
            print(f"Error creating ZIP file: {e}")
            return False


class ExportScheduler:
    """Schedule automatic exports"""

    def __init__(self):
        self.scheduled_exports: List[Dict[str, Any]] = []

    def schedule_export(self, form_id: str, format: str,
                       frequency: str, destination: str) -> str:
        """
        Schedule an automatic export

        Args:
            form_id: Form ID
            format: Export format
            frequency: daily, weekly, monthly
            destination: Export destination path or email

        Returns:
            Schedule ID
        """
        import uuid

        schedule_id = str(uuid.uuid4())
        schedule = {
            "id": schedule_id,
            "form_id": form_id,
            "format": format,
            "frequency": frequency,
            "destination": destination,
            "enabled": True,
            "last_run": None,
        }

        self.scheduled_exports.append(schedule)
        return schedule_id

    def disable_schedule(self, schedule_id: str) -> bool:
        """Disable a scheduled export"""
        for schedule in self.scheduled_exports:
            if schedule["id"] == schedule_id:
                schedule["enabled"] = False
                return True
        return False

    def get_schedules(self, form_id: Optional[str] = None) -> List[Dict[str, Any]]:
        """Get all scheduled exports, optionally filtered by form"""
        if form_id:
            return [s for s in self.scheduled_exports if s["form_id"] == form_id]
        return self.scheduled_exports
